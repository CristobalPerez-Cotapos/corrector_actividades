import os
import subprocess
import openpyxl
import shutil

actividad = "AB"
archivo_a_buscar = "yolanda.py" # Va a ir al repo del alumno y va a seguir 
                                # ingresando a las carpetas hasta encontrar este archivo

archivos_a_agregar = ["api.py", f"test_{actividad}.py"]


def corrector_de_actividades(actividad, archivo_a_buscar):
    carpeta_de_repositorios = actividad
    estudiantes = os.listdir(carpeta_de_repositorios)
    # eliminamos el .git
    if ".git" in estudiantes:
        estudiantes.remove(".git")
    if not os.path.exists("resultados.xlsx"):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("Hoja1")
        wb.save('resultados.xlsx')
    
    wb = openpyxl.load_workbook("resultados.xlsx")
    ws = wb["Hoja1"]

    print("Copiando los tests a la carpeta de los estudiantes")
    estudiantes_paths = {}
    contador_estudiantes = 1
    # Primero copiamos los archivos de test.py a cada carpeta de estudiante
    for estudiante in estudiantes:
        estudiante_dir = os.path.join(carpeta_de_repositorios, estudiante)
        for path, _, f in os.walk(estudiante_dir):
            for file in f:
                if file.lower() == archivo_a_buscar:  # NOMBRE DEL ARCHIVO A BUSCAR
                    for archivo in archivos_a_agregar:
                        agregar_archivo(path, archivo)

                    estudiantes_paths[estudiante] = path

    print("Ya se copiaron los archivos de test.py a cada carpeta de estudiante")

    original_path = os.getcwd()
    for estudiante in estudiantes_paths:
        print("Corrigiendo a", estudiante)
        estudiante_dir = estudiantes_paths[estudiante]
        os.chdir(estudiante_dir)
        subprocess.run(
            f'python3 -m unittest discover -p "test_{actividad}.py" -b 2>&1 | grep "." >> resultados.txt',
            shell=True,
        )
        os.chdir(original_path)

        # leer el archivo de resultados
        final_path = os.path.join(estudiante_dir, "resultados.txt")
        resultados = open(final_path, "r")
        # tomamos la primera linea
        linea = resultados.readline()
        # la linea tiene formato .....F...E.....
        # donde F es falla y E es error
        # Creamos una fila por estudiante y registramos los resultados de cada test
        ws.cell(row=contador_estudiantes, column=1).value = estudiante
        for i in range(len(linea)):
            if linea[i] == "F":
                ws.cell(row=contador_estudiantes, column=i + 2).value = "F"  # Fallido
            elif linea[i] == "E":
                ws.cell(row=contador_estudiantes, column=i + 2).value = "E"  # Error
            elif linea[i] == ".":
                ws.cell(row=contador_estudiantes, column=i + 2).value = "A"  # Aprobado
        contador_estudiantes += 1
        wb.save("resultados.xlsx")


def recolectar_resultados(actividad):
    print("Recolectando resultados")
    # Recorre la carpeta de cada estudiante y copia el archivo de resultados a la carpeta de resultados
    carpeta_de_repositorios = actividad
    estudiantes = os.listdir(carpeta_de_repositorios)
    # eliminamos el .git
    if ".git" in estudiantes:
        estudiantes.remove(".git")
    os.makedirs(f"resultados {actividad}", exist_ok=True)
    for estudiante in estudiantes:
        for path, _, f in os.walk(os.path.join(carpeta_de_repositorios, estudiante)):
            for file in f:
                if file.lower() == "resultados.txt":
                    archivo_a_mover = os.path.join(path, "resultados.txt")
                    destino = os.path.join(
                        f"resultados {actividad}", estudiante + ".txt"
                    )
                    shutil.move(archivo_a_mover, destino)
    print("Resultados recolectados")


def agregar_archivo(path, archivo_a_agregar):
    file_path = os.path.join(path, archivo_a_agregar)
    shutil.copy(archivo_a_agregar, file_path)

if __name__ == "__main__":
    corrector_de_actividades(actividad, archivo_a_buscar)
    recolectar_resultados(actividad)
